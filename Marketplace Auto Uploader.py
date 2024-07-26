import tkinter as tk
import pandas as pd
import os
import time
import threading
import glob
import random
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (ElementClickInterceptedException, NoSuchElementException, TimeoutException, NoSuchWindowException, ElementNotInteractableException)
import traceback
import logging
from tkinter import filedialog, messagebox
import json
from itertools import islice
from openpyxl import load_workbook
from datetime import datetime
from natsort import natsorted, index_natsorted
from selenium.webdriver.common.action_chains import ActionChains

#DARLE SHARE A MI PERFIL LOS UPLOADS QUE HAGA A MARKETPLACE.


#compartir en grupos
#compartir en página (programa aparte, necesita accesar a la pagina primero, luego entrar a listados y repostear - esperar a pagina web para hacer referencia allá?)

# TODO: (low) check for redundancies/inefficiencies.(excepts/exception handling duplication in both parent and child functions/disable-enable inputs and begin button... etc)
# TODO: (Low) folder product ids and excel product ids should be compared in all caps, even if all the data is always entered in all caps.
# TODO: (Low) in input dropdowns, could use a whitespace strip and evaluate choice with dropdown list in all caps, even if all the data is always entered the same as in list.
# BUG: (High) input dropdowns needs to capture ElementNotInteractableException. 
# TODO: instead of time. sleep, can it have a loop to check each two seconds during 10 seconds if the options are displayed? that way, one can make sure that it will wait a reasonable amout of time before identifying and clicking the choice.
# Remove those that can't be uploaded because there's no way of retrieving data for it from anywhere. no product link for example.
# TODO: (done by calling generateproductid/loadvalidatedata on finished upload) should update dataframe after each upload (Uploaded to Site column) so it isn't uploaded again if the program is ran again, or should it just be removed from self.products_to_upload?
# TODO: create a log for valid_df?
#TODO: logging validation issues, but not products that could not be uploaded while in process of uploading. should be added? I need to know to fix data or program.
#TODO: load and validate data and generate product id list needs migration at the beginning of start_upload_process
# function needs migration to start_upload_process because it increases reliability. when program is run, it gets the data to upload on load, should do on begin press
# which helps to add new data while the program is open, and the program taking in consideration that new data on begin, rather than needing to restart the program.
# could also just be to remove call on folder/excel selection, might be easier than migrating everything. check.
#TODO: create a new program to get all the product ids from folder and from excel and check which product ids are missing from each.
#TODO GPT CANT GIVE TITLES LONGER THAN 97 CHARS

# Get the directory of the script
script_dir = os.path.dirname(os.path.realpath(__file__))

# Set the current working directory to the script's directory
os.chdir(script_dir)

validation_log_file = 'needs_user_action_before_upload.log'

# Check if the validation log file exists and delete it
if os.path.exists(validation_log_file):
    os.remove(validation_log_file)

logging.basicConfig(filename='marketplace_uploader.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Separate logger for validation issues (User report)
validation_logger = logging.getLogger('needs_user_action_before_upload')
validation_file_handler = logging.FileHandler(validation_log_file, mode='a')
validation_formatter = logging.Formatter('%(asctime)s - %(message)s')
validation_file_handler.setFormatter(validation_formatter)
validation_logger.addHandler(validation_file_handler)
validation_logger.setLevel(logging.ERROR)

class MarketplaceUploader:

# On load:
# Initialize GUI and load URL's
    def __init__(self):
        """Initializes the MarketplaceUploader with GUI components and configurations."""
        self.script_dir = os.path.dirname(os.path.realpath(__file__))
        os.chdir(self.script_dir)

        self.root = tk.Tk()
        self.root.title("Facebook auto Marketplace uploader")
        self._setup_gui()
        self.driver = None  
        self.product_folder_selected = False 
        self.excel_file_path = None
        self.excel_sheet_name = None
        self.config_path = 'config.json'
        self._load_config()

        self.products_to_upload_df = None  # This will store the filtered and matching data
        self.upload_limit = 500  # Maximum number of products to upload 

        # Load and validate the DataFrame during initialization
        self.valid_df = None  
        if self.excel_file_path and self.excel_sheet_name:
            self.load_and_validate_data()

        # Check for valid product folder on program load
        self.folders_product_ids = None
        self.product_folder_path = self.check_and_load_product_folder()
        if self.product_folder_path:
            self.generate_product_ids_list()

        # Initialize a threading event for stopping threads
        self.stop_event = threading.Event()
        
        # Initialize the pause event
        self.pause_event = threading.Event()
        self.pause_event.set()  # Start with the event set (i.e., not paused)

    def run(self):
        """Runs the main loop of the Tkinter application."""
        self.root.mainloop()

    def _setup_gui(self):
        """Sets up the GUI components for the application."""
        # Check and read URLs from files
        website_url, marketplace_url = self._read_urls()

        # Create and place widgets for website and marketplace links, email, and password
        tk.Label(self.root, text="Enter website link here:").grid(row=0, column=0)
        self.link_entry = self._create_entry(website_url, row=0)

        tk.Label(self.root, text="Enter website Marketplace link here:").grid(row=1, column=0)
        self.marketplace_link_entry = self._create_entry(marketplace_url, row=1)

        tk.Label(self.root, text="Email:").grid(row=2, column=0)
        self.email_entry = self._create_entry(row=2)

        tk.Label(self.root, text="Password:").grid(row=3, column=0)
        self.password_entry = self._create_entry(row=3, show="*")

        # Create and place Begin button and login status label
        self.begin_button = tk.Button(self.root, text="Begin", command=self.start_upload_process)  
        self.begin_button.grid(row=6, column=1)
        # Initially disable the Begin button until conditions are met
        self.begin_button.config(state=tk.DISABLED) 

        self.login_status_label = tk.Label(self.root, text="")
        self.login_status_label.grid(row=7, column=1)

        # Excel file selection button and label
        self.excel_label = tk.Label(self.root, text="None selected")
        self.excel_label.grid(row=4, column=1)
        
        self.select_excel_button = tk.Button(self.root, text="Select Excel File", command=self.select_excel_file)
        self.select_excel_button.grid(row=4, column=0)

        # Product folder selection button and label
        self.folder_label = tk.Label(self.root, text="None selected")
        self.folder_label.grid(row=5, column=1)

        self.select_product_folder_button = tk.Button(self.root, text="Select Product Folder", command=self.select_product_folder)
        self.select_product_folder_button.grid(row=5, column=0)

        # Add a Pause button and Resume button side by side
        self.pause_button = tk.Button(self.root, text="Pause", command=self.toggle_pause, state=tk.DISABLED)
        self.pause_button.grid(row=8, column=0)

        self.resume_button = tk.Button(self.root, text="Resume", command=self.toggle_pause, state=tk.DISABLED)
        self.resume_button.grid(row=8, column=1)

        # Add a status label to show the current status
        self.status_label = tk.Label(self.root, text="Status: Idle")
        self.status_label.grid(row=8, column=2)

        # Bind <KeyRelease> events to update the Begin button state
        self.link_entry.bind('<KeyRelease>', lambda e: self.check_conditions_and_update_begin_button())
        self.marketplace_link_entry.bind('<KeyRelease>', lambda e: self.check_conditions_and_update_begin_button())
        self.email_entry.bind('<KeyRelease>', lambda e: self.check_conditions_and_update_begin_button())
        self.password_entry.bind('<KeyRelease>', lambda e: self.check_conditions_and_update_begin_button())

        self.root.bind('<Return>', self.on_enter_pressed)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        self.check_conditions_and_update_begin_button()


    def toggle_pause(self):
        if self.pause_event.is_set():
            # Currently running, so we pause
            self.pause_event.clear()
            self.pause_button.config(state=tk.DISABLED)
            self.resume_button.config(state=tk.NORMAL)
            self.update_label("Upload paused. Press resume to continue.")
            self.status_label.config(text="Status: Paused")
        else:
            # Currently paused, so we resume
            self.pause_event.set()
            self.pause_button.config(state=tk.NORMAL)
            self.resume_button.config(state=tk.DISABLED)
            self.update_label("Upload resumed.")
            self.status_label.config(text="Status: Uploading")

    def update_label(self, message):
        self.login_status_label.config(text=message)

    def wait_for_resume(self):
        """Waits for the pause event to be set, effectively pausing the thread if cleared."""
        while not self.pause_event.is_set():
            time.sleep(1)  # Sleep briefly to wait for the pause event to be set

    def _read_urls(self):
        """Reads website and marketplace URLs from their respective files."""
        website_url, marketplace_url = "", ""
        if os.path.exists("website_url.txt"):
            with open("website_url.txt", "r") as file:
                website_url = file.read().strip()
        if os.path.exists("website_marketplace_link.txt"):
            with open("website_marketplace_link.txt", "r") as file:
                marketplace_url = file.read().strip()
        return website_url, marketplace_url

    def _create_entry(self, default_text="", row=0, show=None):
        """Creates an Entry widget with default text and specified properties."""
        entry = tk.Entry(self.root, width=50, show=show)
        entry.insert(0, default_text)
        entry.grid(row=row, column=1)
        return entry


# Loads folders and store them if they have valid data
    def check_and_load_product_folder(self):
        """Checks for a valid product folder from a saved JSON file on program load.
        
        If the JSON file exists and contains a valid path with product folders, returns the path.
        If no valid folder is found, deletes the existing JSON file.
        """
        json_path = 'product_folder_path.json'
        if os.path.exists(json_path):
            with open(json_path, 'r') as file:
                data = json.load(file)
                product_folder_path = data.get('path', '')
                if self.is_valid_product_folder(product_folder_path):
                    logging.info(f"Valid product folder found: {product_folder_path}")
                    self.folder_label.config(text=f'Folder Selected: {os.path.basename(product_folder_path)}')
                    self.check_conditions_and_update_begin_button()
                    return product_folder_path
                else:
                    logging.info("No valid product folders found. Deleting JSON file.")
                    os.remove(json_path)
        return ""

    def is_valid_product_folder(self, path):
        """Checks if the provided path has at least one valid product folder.
        
        Args:
            path (str): The path to check for valid product folders.
        
        Returns:
            bool: True if valid product folders are found, False otherwise.
        """
        try:
            for folder in os.listdir(path):
                if not folder.startswith('-') and os.path.isdir(os.path.join(path, folder)):
                    return True
        except Exception as e:
            logging.error(f"Error checking valid product folders: {e}")
        return False

    def select_product_folder(self):
        """Prompts the user to select a product folder and validates the selection.

        If the selected folder is valid, saves its path to a JSON file.
        If invalid, shows an error message and does not save the path.
        """
        folder_path = filedialog.askdirectory(title="Select Product Folder")
        if folder_path and self.is_valid_product_folder(folder_path):
            self.save_product_folder_path(folder_path)
            logging.info(f"Product folder selected and saved: {folder_path}")
            messagebox.showinfo("Folder Selection", "Valid product folder selected.")
            self.folder_label.config(text=f'Folder Selected: {os.path.basename(folder_path)}')
            
            self.product_folder_path = self.check_and_load_product_folder()
            self.generate_product_ids_list()
            
            self.check_conditions_and_update_begin_button()
        else:
            messagebox.showerror("Folder Selection Error", "Invalid folder. No product folders inside. \n\nKeeping previously selected folder path.")

    def save_product_folder_path(self, path):
        """Saves the selected product folder path to a JSON file.

        Args:
            path (str): The path to save.
        """
        data = {'path': path}
        with open('product_folder_path.json', 'w') as file:
            json.dump(data, file)
        logging.info("Product folder path saved to JSON.")

    def generate_product_ids_list(self):
        """Generates a list of product IDs from the selected product folder.

        Returns:
            list: A list of product IDs.
        """
        product_ids = []
        if self.product_folder_path:
            try:
                for folder in os.listdir(self.product_folder_path):
                    folder_path = os.path.join(self.product_folder_path, folder)
                    if not folder.startswith('-') and os.path.isdir(folder_path):
                        # Check for pictures in the folder
                        total_pictures = 0
                        for file in os.listdir(folder_path):
                            if file.lower().endswith(('.jpg', '.jpeg', '.png')):
                                total_pictures += 1
                        
                        product_id = folder.split()[0]  # Assuming the product ID is the first part before a space
                        if 0 < total_pictures <= 10:
                            product_ids.append(product_id)
                            logging.info(f"Found valid folder with product ID: {product_id} with {total_pictures} pictures.")  # Print each found product ID
                        else:
                            logging.error(f"Found folder with product ID: {product_id} but contains {total_pictures} pictures. Folder skipped for upload.")
                            validation_logger.error(f"Folder with product ID: {product_id} requires updating: contains {total_pictures} images")

                validation_logger.error(f"---------------------------------------------------------------------------")
                self.folders_product_ids = product_ids
                print(f"Valid folders in selected folder: {len(self.folders_product_ids)}")
            except Exception as e:
                logging.error(f"Error generating product IDs list: {e}")
                print(f"Error generating product IDs list: {e}")


# Loads excel and store rows if they have valid data
    def select_excel_file(self):
        file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.excel_file_path = file_path
            self._select_sheet(file_path)
            
    def _select_sheet(self, file_path):
        xl = pd.ExcelFile(file_path)
        sheets = xl.sheet_names
        
        def on_sheet_selected(sheet_name):
            self.excel_sheet_name = sheet_name
            self._save_config()
            self._update_excel_label()  # Update the label after selecting a sheet
            self.load_and_validate_data()
            self.check_conditions_and_update_begin_button()
            sheet_window.destroy()

        # Create a new Toplevel window
        sheet_window = tk.Toplevel(self.root)
        sheet_window.title("Select a Sheet")
        button_width = 20

        for sheet in sheets:
            btn = tk.Button(sheet_window, text=sheet, width=button_width, anchor="w",
                            command=lambda s=sheet: on_sheet_selected(s))
            btn.pack(pady=2)

        sheet_window.update_idletasks()
        window_width = sheet_window.winfo_width()
        window_height = sheet_window.winfo_height()
        screen_width = sheet_window.winfo_screenwidth()
        screen_height = sheet_window.winfo_screenheight()
        center_x = int(screen_width/2 - window_width/2)
        center_y = int(screen_height/2 - window_height/2)
        sheet_window.geometry(f'+{center_x}+{center_y}')

    def _update_excel_label(self):
        if self.excel_file_path and self.excel_sheet_name:
            file_name = os.path.basename(self.excel_file_path)
            self.excel_label.config(text=f"File: {file_name}, Sheet: {self.excel_sheet_name}")
        else:
            self.excel_label.config(text="No Excel file selected")

    def _save_config(self):
        config = {
            'excel_file_path': self.excel_file_path,
            'excel_sheet_name': self.excel_sheet_name
        }
        with open(self.config_path, 'w') as f:
            json.dump(config, f)
    
    def _load_config(self):
        if os.path.exists(self.config_path):
            with open(self.config_path, 'r') as f:
                config = json.load(f)
                self.excel_file_path = config.get('excel_file_path')
                self.excel_sheet_name = config.get('excel_sheet_name')
                self._update_excel_label()  # Update the label with loaded values

# Checks if excel has valid products to upload
    def load_and_validate_data(self):
        """Loads the DataFrame from the Excel file and validates it."""
        try:
            df = pd.read_excel(self.excel_file_path, sheet_name=self.excel_sheet_name)
            valid_count = 0  # Initialize a counter for valid rows

            # Convert 'To Sell Date' to datetime format for the entire column, ensuring errors are coerced to NaT
            df['To Sell After'] = pd.to_datetime(df['To Sell After'], errors='coerce', dayfirst=True)

            today = pd.Timestamp('today').normalize()  # Get today's date for comparison

            # Filtering criteria
            conditions = (
                (df['Damaged'] != 'YES') &
                (df['Personal'] != 'YES') &
                (df['Cancelled Order'] != 'YES') &
                (df['Uploaded to Site'] != 'YES') &
                (df['Sold'] != 'YES') &
                df['To Sell After'].notna() & (df['To Sell After'].dt.floor('D') <= today) &  # Correctly compare dates
                df['Product Price After IVU'].notna() &
                df['Category'].notna() &
                df['Condition'].notna() &
                df['Product Description'].notna() &
                df['Product Tags'].notna() &
                df['Product ID'].notna() &
                df['Rack ID'].notna() 
            )

            # Apply filters
            valid_df = df[conditions]

            # Logging reasons for invalid rows
            for index, row in df.iterrows():
                if row['Damaged'] != 'YES' and row['Personal'] != 'YES' and row['Cancelled Order'] != 'YES' and row['Uploaded to Site'] != 'YES' and row['Sold'] != 'YES':
                    # These rows are potentially eligible for upload; now check for missing data
                    issues = []
                    if 'Título:' not in str(row['Product Description']): issues.append('Título: missing')
                    if pd.isna(row['Product Price After IVU']): issues.append('Product Price After IVU missing')
                    if pd.isna(row['Category']): issues.append('Category missing')
                    if pd.isna(row['Condition']): issues.append('Condition missing')
                    if pd.isna(row['Product Description']): issues.append('Description missing')
                    if pd.isna(row['Product Tags']): issues.append('Product Tags missing')
                    if pd.isna(row['Product ID']): issues.append('Product ID missing')
                    if pd.isna(row['Rack ID']): issues.append('Rack ID missing')
                    if pd.isna(row['To Sell After']): issues.append('To Sell After date missing')

                    if issues:
                        validation_logger.error(f"Row {index + 1} requires updating: {', '.join(issues)}")
                    else:
                        valid_count += 1

            validation_logger.error(f"---------------------------------------------------------------------------")

            # Store valid rows to upload
            self.valid_df = valid_df

            # Checks the rows stored in the dataframe
            print(f"Rows requiring data input: {valid_count}")
            print(f"Rows with all required data for upload: {len(self.valid_df)}")

        except Exception as e:
            print(f"Error loading or validating Excel file: {e}")
            self.valid_df = None


# Check if all inputs have been made to enable/disable the Begin button 
    def check_conditions_and_update_begin_button(self):
        # Check if all text entries have content and both the Excel file and product folder have been selected
        if (self.link_entry.get() and self.marketplace_link_entry.get() and
            self.email_entry.get() and self.password_entry.get() and
            self.excel_file_path and self.excel_sheet_name and self.product_folder_path):
            self.begin_button.config(state=tk.NORMAL)  # Enable the "Begin" button
        else:
            self.begin_button.config(state=tk.DISABLED)  # Disable the "Begin" button
    
# Updates status label whenever a change is made in the program.
    def update_label(self, message):
        self.login_status_label.config(text=message)


# On Begin button pressed:
# Begins 
    def on_enter_pressed(self, event=None):
        """Method to be called when the Enter key is pressed.
        
        It checks if the 'Begin' button is enabled and if so, starts the login process.
        """
        if self.begin_button['state'] == tk.NORMAL:
            self.start_upload_process()

    def disable_inputs(self):
        self.link_entry.config(state='disabled')
        self.marketplace_link_entry.config(state='disabled')
        self.email_entry.config(state='disabled')
        self.password_entry.config(state='disabled')
        self.select_excel_button.config(state='disabled')  
        self.select_product_folder_button.config(state='disabled')  
        
    def enable_inputs(self):
        self.link_entry.config(state='normal')
        self.marketplace_link_entry.config(state='normal')
        self.email_entry.config(state='normal')
        self.password_entry.config(state='normal')
        self.select_excel_button.config(state='normal')
        self.select_product_folder_button.config(state='normal')

# Saves the Login URL and Marketplace URL entered for future uses/program execution
    def save_links(self):
        """
        Save the website URL and marketplace URL in their respective text files.
        """
        with open("website_url.txt", "w") as file:
            file.write(self.link_entry.get())

        with open("website_marketplace_link.txt", "w") as file:
            file.write(self.marketplace_link_entry.get())

# Checks if there's a match between the picture folder's product ID's and Excel's product ID's before attempting to access the web and upload products
    def start_upload_process(self):
        """Starts the upload process if there is valid data to upload."""

        self.load_and_validate_data() 
        self.generate_product_ids_list() 

        self.disable_inputs()  # Disable inputs at the start
        self.begin_button.config(state=tk.DISABLED)  # Disable the "Begin" button
        self.pause_button.config(state=tk.NORMAL)  # Enable the "Pause" button
        self.resume_button.config(state=tk.DISABLED)  # Disable the "Resume" button

        if self.valid_df is not None and not self.valid_df.empty:
            # Extract and sort Product IDs from the valid DataFrame
            excel_product_ids = natsorted(self.valid_df['Product ID'].tolist())
            
            # Sort Product IDs from the folders product ids
            folder_product_ids = natsorted(self.folders_product_ids)
            
            # Find matching Product IDs
            matching_product_ids = [pid for pid in folder_product_ids if pid in excel_product_ids]

            if not matching_product_ids:
                # No matching Product IDs found
                messagebox.showinfo("Upload Process", "No folder product IDs found in the Excel product IDs.")
                print(f"Folder Product IDs: {folder_product_ids}")
                print(f"Excel Product IDs from valid DataFrame: {excel_product_ids}")
                print(f"Folder Product ID count: {len(folder_product_ids)}")
                print(f"Valid DataFrame row count: {len(self.valid_df)}")
            else:
                # Print matching and all folder and Excel product IDs, and the counts
                print(f"Folder Product IDs: {folder_product_ids}")
                print(f"Excel Product IDs from valid DataFrame: {excel_product_ids}")
                print(f"Matching Product IDs: {matching_product_ids}")

                products_to_upload = self.valid_df[self.valid_df['Product ID'].isin(matching_product_ids)]

                # Sort products_to_upload by Product ID in natural order
                order = index_natsorted(products_to_upload['Product ID'])
                self.products_to_upload_df = products_to_upload.iloc[order]


                print(f"Products to upload: {len(matching_product_ids)} out of {len(self.valid_df)} products available in the Excel to upload.")
                print("")
                print("Note: To increase the total products to upload, the program needs to find more matches, which requires the following:")
                print("1) Add the product folder with the ID that's present in the Excel product ID column and make the pictures for that product be between 1 and 10.")
                print("2) Add the product ID's in the Excel to match the product ids from both the excel and the folders.")
                print("")
                print(f"Total products to upload: {len(self.products_to_upload_df)}")


                # After determining which products to upload, print the first set to be uploaded
                upload_preview_limit = self.upload_limit  # or any number you want to preview
                upload_preview = self.products_to_upload_df.head(upload_preview_limit)
                print(f"First {upload_preview_limit} products scheduled for upload:")
                print(upload_preview[['Product ID', 'Product Description']])

            # Access the web after at least one matching product id(valid excel data/folders with images)
            self.login_thread = threading.Thread(target=self.attempt_login, daemon=True)
            self.login_thread.start()

        else:
            print("No valid data to upload. Please check the criteria and the Excel file.")
            self.on_process_complete()  # Ensure UI is reset if there's no valid data


# Attempts to log in to URL
    def attempt_login(self, event=None):
        """
        Retrieves email, password, and website URL from the Tkinter GUI and attempts to log in using Selenium in headless mode.
        Includes a random sleep timer between 3 to 5 seconds after logging in.
        Updates the GUI with the login status.
        Saves the website URL in website_url.txt.
        """
        while not self.stop_event.is_set():

            logging.info("Attempting to log in.")

            self.save_links()

            # Set Chrome options to disable notifications
            chrome_options = Options()
            chrome_options.add_argument("--disable-notifications")
            #chrome_options.add_argument("--headless")

            # Initialize the WebDriver with the specified options
            self.driver = webdriver.Chrome(options=chrome_options)  # Store self.driver as an instance variable

            try:
                # Open the login page
                self.driver.get(self.link_entry.get())
                print("Opened the login page.")
                logging.info("Opened the login page.")
                self.update_label("Opened the login page.")
                time.sleep(random.uniform(1.0, 2.0))

                # Find and fill the email input field
                email_field = self.driver.find_element(By.NAME, "email")
                email_field.send_keys(self.email_entry.get())
                print("Email entered.")
                self.update_label("Email entered.")
                time.sleep(random.uniform(1.0, 2.0))

                # Find and fill the password input field
                password_field = self.driver.find_element(By.NAME, "pass")
                password_field.send_keys(self.password_entry.get())
                print("Password entered.")
                self.update_label("Password entered.")
                time.sleep(random.uniform(1.0, 2.0))

                # Find and click the login button
                login_button = self.driver.find_element(By.NAME, "login")
                login_button.click()
                print("Login button clicked.")
                self.update_label("Login button clicked.")

                # Initialize a variable to track the total wait time
                total_wait_time = 0

                while total_wait_time < 20:  # Check for up to 20 seconds
                    # Retrieve the class attribute of the <body> tag
                    body_class = self.driver.find_element(By.TAG_NAME, "body").get_attribute("class")

                    # Check if "UIPage_LoggedOut" is not in the class attribute
                    if "UIPage_LoggedOut" not in body_class:
                        self.update_label("Login successful! Moving to marketplace...")
                        print("Login successful! Moving to marketplace...")
                        self.update_label("Login successful! Moving to marketplace...")
                        self.access_marketplace()  
                        return  # Exit the function after access marketplace finishes execution

                    # Wait for 2 seconds before checking again
                    time.sleep(2)
                    total_wait_time += 2

                else:  # This block executes if the loop completes without breaking (i.e., login failed)
                    self.update_label("Login failed.")
                    print("Login failed.")
            except NoSuchElementException as e:
                self.update_label(f"Login element not found.")
                print(f"Login element not found: {e}")
                logging.error(f"Login element not found: {e}")

            except TimeoutException as e:
                self.update_label(f"Page load timeout.")
                print(f"Page load timeout: {e}")
                logging.error(f"Page load timeout: {e}")

            except NoSuchWindowException as e:
                logging.error("Browser window was closed unexpectedly.")
                self.update_label("Browser window closed unexpectedly.")
                print("Browser window closed unexpectedly.")

            except Exception as e:
                self.update_label(f"Unexpected error during login.")
                print(f"Unexpected error during login: {e}")
                logging.error(f"Unexpected error during login: {e}")

            finally:
                self.cleanup_resources()  # Ensure the browser is closed if login fails
                # Schedule the check_conditions_and_update_begin_button to run on the main thread
                self.root.after(0, self.check_conditions_and_update_begin_button)
                self.on_process_complete()
            break  # Example: break the loop if stop_event is set

# Attempts to access the Marketplace URL
    def access_marketplace(self):
        """
        Perform tasks after a successful login, iterating through each product for upload.
        """
        time.sleep(random.uniform(1.0, 2.0))
        uploaded_count = 0 

        print("Starting product upload process...")
        self.update_label("Starting product upload process...")
        self.status_label.config(text="Status: Uploading")
        logging.info("Starting product upload process...")

        try:
            # Use enumerate to get both index and product, and limit the number of iterations with islice
            for index, product in islice(self.products_to_upload_df.iterrows(), self.upload_limit):
                self.wait_for_resume()  # Wait if paused before starting each upload
                # Navigate to the marketplace link at the start of each iteration
                self.driver.get(self.marketplace_link_entry.get())
                print("Navigating to marketplace...")
                self.update_label("Navigating to marketplace...")
                logging.info("Navigating to marketplace...")

                try:
                    # Wait up to 20 seconds for the marketplace to be ready
                    WebDriverWait(self.driver, 20).until(
                        EC.presence_of_all_elements_located((By.XPATH, "//*[contains(text(), 'Marketplace') or contains(text(), 'Item For Sale') or contains(text(), 'Listing to Marketplace')]"))
                    )
                    print("Found Marketplace ready for product upload.")
                    
                    # Upload the product using the extracted details
                    product_id = product['Product ID']
                    rack_id = product['Rack ID']
                    price = product['Product Price After IVU']
                    category = product['Category']
                    condition = product['Condition']
                    description = product['Product Description']
                    tags = product['Product Tags']
                    title = self.extract_title(description)
                    description = self.adjust_description(description)

                    # Upload the product using the extracted details
                    success = self.input_product_data(product_id, rack_id, title, price, category, condition, description, tags)

                    if success:
                        uploaded_count += 1
                        print(f"Product {product_id} uploaded successfully. Total uploaded: {uploaded_count}")
                        self.update_label(f"Product {product_id} uploaded successfully. Total uploaded: {uploaded_count}")
                        logging.info(f"Product {product_id} uploaded successfully. Total uploaded: {uploaded_count}")
                        self.update_uploaded_status_in_excel(product_id)
                        self.create_uploaded_successful_txt(product_id)
                    else:
                        print(f"Failed to upload product {product_id}.")
                        self.update_label(f"Failed to upload product {product_id}.")
                        logging.info(f"Failed to upload product {product_id}.")
                        
                except TimeoutException:
                    print("Marketplace not found within 20 seconds.")
                    self.update_label("Marketplace not found within 20 seconds.")
                    logging.warning("Marketplace not found within 20 seconds.")
                    return  # Exit the function due to marketplace access issues

                # Check pause state before next iteration
                self.wait_for_resume()

                # If the uploaded count reaches the limit, break the for loop
                if uploaded_count >= self.upload_limit:
                    print(f"Reached upload limit of {self.upload_limit} products.")
                    break

            print(f"Product upload process completed. Total products uploaded: {uploaded_count} out of {len(self.products_to_upload_df)}")
            self.update_label(f"Product upload process completed.")
            self.status_label.config(text="Status: Idle")
            logging.info(f"Product upload process completed. Total products uploaded: {uploaded_count}")
            messagebox.showinfo("Upload Process", f"Product upload process completed. \n\nTotal products uploaded: {uploaded_count}")

        except NoSuchWindowException as e:
            logging.error("Browser window was closed unexpectedly.")
            self.update_label("Browser window closed unexpectedly.")
            print("Browser window closed unexpectedly.")
        except NoSuchElementException:
            self.update_label("Marketplace page structure might have changed.")
            print("Marketplace page structure might have changed.")
            logging.error(f"Marketplace page structure might have changed.")
        except Exception as e:
            self.update_label(f"Unexpected error while accessing marketplace.")
            print(f"Unexpected error while accessing marketplace: {e}")
            logging.error(f"Unexpected error while accessing marketplace: {e}")
        finally:
            self.cleanup_resources()
            self.on_process_complete()


# Attempts to input pictures and data to Marketplace
    def input_product_data(self, product_id, rack_id, title, price, category, condition, description, tags):

        print("Starting product upload...")
        self.update_label("Starting product upload...")
        logging.info("Starting product upload...")

        print(f"Product to upload: {product_id}")

        try:
            time.sleep(random.uniform(1.0, 2.0))
            
            # Image upload
            self.wait_for_resume()
            self.upload_images(product_id)
            
            # Title input
            self.wait_for_resume()
            self.input_title(title)
            
            # Price input
            self.wait_for_resume()
            self.input_price(price)
            
            # Category selection
            self.wait_for_resume()
            self.input_category(category)

            # Condition selection
            self.wait_for_resume()
            self.input_condition(condition)
            
            # Description input
            self.wait_for_resume()
            self.input_description(description)

            # Product Tags input
            self.wait_for_resume()
            self.input_product_tags(tags)

            # Description input
            self.wait_for_resume()
            self.input_sku(product_id, rack_id)
            
            # Clicking the "Next" button
            self.wait_for_resume()
            self.click_next_button()
            
            # Selecting all the checkboxes
            self.wait_for_resume()
            self.select_checkboxes()

            # Clicking the "Publish" button
            self.wait_for_resume()
            self.click_publish_button()
            
            time.sleep(5)
            # See if successfully published.
            # Check for URL redirection after clicking "Publish"
            expected_url = "https://www.facebook.com/marketplace/you/selling"
            start_time = time.time()
            while time.time() - start_time <= 30:  # Check for up to 20 seconds
                current_url = self.driver.current_url
                if current_url == expected_url:
                    print("Redirected to expected URL after publish.")
                    logging.info("Redirected to expected URL after publish.")
                    self.update_label("Redirected to expected URL after publish.")
                    return True  # Successful redirection, hence successful upload
                time.sleep(2)  # Wait for 2 seconds before checking again

            print("Product seems to not have been uploaded. Did not redirect to expected URL after publish.")
            logging.error("Product seems to not have been uploaded. Did not redirect to expected URL after publish.")
            self.update_label("Product seems to not have been uploaded. Did not redirect to expected URL after publish.")
            return False  # Redirection did not occur as expected
        except NoSuchWindowException as e:
            raise
        except Exception as e:
            print(f"Error occurred: {e}")
            self.update_label(f"Error occurred during product upload: {e}")
            logging.error(f"Error occurred during product upload: {e}")
            traceback.print_exc()
            return False  # Return False to indicate failure



    def upload_images(self, product_id):
        """Uploads images for a given product ID from the selected product folder.

        Args:
            product_id (str): The product ID whose images are to be uploaded.
        """

        try:        
            # Ensure there's a selected product folder path
            if not self.product_folder_path:
                logging.error("No product folder path selected.")
                raise Exception("No product folder path selected.")

            # Find the folder that starts with the product ID
            product_folder = None
            for folder in os.listdir(self.product_folder_path):
                if folder.startswith(product_id) and os.path.isdir(os.path.join(self.product_folder_path, folder)):
                    product_folder = os.path.join(self.product_folder_path, folder)
                    break

            if not product_folder:
                logging.error(f"No folder for product ID {product_id} found.")
                raise Exception(f"No folder for product ID {product_id} found.")

            product_folder = os.path.normpath(product_folder)

            print("Checking folder:", product_folder)
            all_files = os.listdir(product_folder)
            print("Files found:", all_files)

            # Create a list of full paths for the .jpg files to be uploaded
            images_to_upload = [os.path.join(product_folder, file) for file in all_files if file.lower().endswith('.jpg')][:10]
            print("Images to upload (with absolute paths):", images_to_upload)

            if not images_to_upload:
                logging.info(f"No images to upload in folder {product_folder}.")
                raise Exception(f"No images to upload in folder {product_folder}.")

            # Simulate uploading images and provide feedback
            logging.info(f"Uploading images for product ID {product_id} from {product_folder}")
            print(f"Uploading images for product ID {product_id} from {product_folder}")
            self.update_label(f"Uploading images for {product_id}...")

            # Find and click the element that triggers the file input (e.g., "Add Photos")
            print("Clicked upload trigger.")
            logging.info("Clicked upload trigger.")
            self.update_label("Clicked upload trigger.")
            
            # Allow time for the file input to become available if needed
            time.sleep(2)  # Adjust sleep time based on how quickly the input becomes available
            
            # Assuming the file input is now interactable, proceed with uploading
            file_input_xpath = "//input[@type='file']"
            file_input = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, file_input_xpath))
            )
            file_paths = '\n'.join(images_to_upload)
            file_input.send_keys(file_paths)
            
            print("Images uploaded successfully.")
            logging.info(f"Images for product ID {product_id} uploaded successfully.")
            self.update_label(f"Images for {product_id} uploaded successfully.")
        except NoSuchWindowException as e:
            raise  # Reraise the exception to signal the failure
        except Exception as e:
            print(f"Error while uploading images: {e}")
            logging.error(f"Error in uploading images for product ID {product_id}: {e}")
            self.update_label(f"Error uploading images for {product_id}. See log for details.")
            traceback.print_exc()
            raise  # Reraise the exception to signal the failure

        time.sleep(random.uniform(1.0, 2.0))

    def input_title(self, title):
        if not title:
            raise Exception("No title provided.")

        print(f"Inputting title: {title}")
        self.update_label("Inputting title.")
        logging.error("Inputting title.")

        # Input Title
        try:
            # Wait for the title input field to be available
            wait = WebDriverWait (self.driver, 10) 
            # Adjusted XPath to target the specific title input field
            title_input_xpath = "//label[@aria-label='Title']//input[@type='text']"
            title_input = wait.until(EC.presence_of_element_located((By.XPATH, title_input_xpath)))
            
            # Slice the title to the first 97 characters
            limited_title = title[:97]

            # Send the limited title to the input field
            title_input.send_keys(limited_title)

            print(f"Title {title} input successful.")
            logging.info(f"Title '{title}' input successful.")
            self.update_label(f"Title input successful.")
        except NoSuchWindowException as e:
            raise  # Reraise the exception to signal the failure
        except Exception as e:
            print(f"Error while inputting title data: {e}")
            logging.error(f"Error while inputting title data: {e}")
            self.update_label(f"Error while inputting title data.")
            traceback.print_exc()  # This will print the full traceback
            raise  # Reraise the exception to signal the failure

        time.sleep(random.uniform(1.0, 2.0))
    
    def extract_title(self, description):
        title_start = description.find("Título:") + 7
        if title_start > 6:
            title_end = description.find('\n', title_start)
            return description[title_start:title_end].strip() if title_end != -1 else description[title_start:].strip()
        return None

    def input_price(self, price):
        if price is None: 
            raise Exception("No price provided.")

        print(f"Inputting price: {price}")
        logging.info(f"Inputting price.")
        self.update_label(f"Inputting price.")
            # Input Price
        try:
            # Initialize WebDriverWait with self.driver
            wait = WebDriverWait(self.driver, 10)  # Adjust the timeout as needed

            # Proceed with finding the price input field and other operations
            price_input_xpath = "//label[@aria-label='Price']//input[@type='text']"
            price_input = wait.until(EC.presence_of_element_located((By.XPATH, price_input_xpath)))

            # Convert price to string if it's not already, to use string methods
            price_str = str(price)

            # Remove the dollar sign if present
            if price_str.startswith("$"):
                price_str = price_str[1:]

            rounded_price = round(float(price_str))

            # Send the rounded integer as a string to price_input
            price_input.send_keys(str(rounded_price))

            print(f"Price {rounded_price} input successful.")
            logging.info(f"Price {rounded_price} input successful.")
            self.update_label(f"Price input successful.")
        except NoSuchWindowException as e:
            raise  # Reraise the exception to signal the failure
        except Exception as e:
            print(f"Error while inputting price data: {e}")
            logging.error(f"Error while inputting price data: {e}")
            self.update_label(f"Error while inputting price data.")
            traceback.print_exc()  # This will print the full traceback
            raise  # Reraise the exception to signal the failure

        time.sleep(random.uniform(1.0, 2.0))

    def input_category(self, category):
        if not category:
            raise Exception("No category provided.")

        print(f"Selecting category: {category}")
        logging.error(f"Selecting category.")
        self.update_label(f"Selecting category.")
        # Input Category
        try:
            # Click the Category dropdown to display options
            category_dropdown_xpath = "//label[@aria-label='Category']"
            category_dropdown = self.driver.find_element(By.XPATH, category_dropdown_xpath)

            # Scroll the element into view
            self.driver.execute_script("arguments[0].scrollIntoView(true);", category_dropdown)
            time.sleep(2)  # Wait for any overlay to disappear

            # Try clicking the dropdown
            try:
                category_dropdown.click()
            except ElementClickInterceptedException:
                # Use JavaScript click as a fallback
                self.driver.execute_script("arguments[0].click();", category_dropdown)

            time.sleep(5)  # Wait for options to be displayed

            # Find all choices that match the category text exactly
            category_choice_xpath = f"//div[@data-visualcompletion='ignore-dynamic']//span[text()='{category}']"
            category_choices = self.driver.find_elements(By.XPATH, category_choice_xpath)  # Note the plural form here

            exact_match_found = False
            if category_choices:  # Ensure there are choices available to iterate
                for choice in category_choices:
                    # Since you're using exact text match with XPath, additional text comparison might not be necessary
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", choice)
                    time.sleep(1)  # Ensure scrolling effect is complete
                    choice.click()
                    exact_match_found = True
                    print("Exact category choice clicked")
                    break

            if not exact_match_found:
                print(f"Exact category '{category}' not found. Please check the available options.")
                logging.error(f"Exact category '{category}' not found.")


            print(f"Category choice '{category}' input successful.")
            logging.info(f"Category choice '{category}' input successful.")
            self.update_label(f"Category choice input successful.")

            time.sleep(5)  # Wait for the selection to be processed
        except NoSuchWindowException as e:
            raise  # Reraise the exception to signal the failure
        except Exception as e:
            print(f"Error while selecting category: {e}")
            logging.error(f"Error while selecting category: {e}")
            self.update_label(f"Error while selecting category.")
            traceback.print_exc()
            raise  # Reraise the exception to signal the failure

        time.sleep(random.uniform(1.0, 2.0))

    def input_condition(self, condition):
        if not condition:
            raise Exception("No condition provided.")

        print(f"Selecting condition: {condition}")
        logging.error(f"Selecting condition.")
        self.update_label(f"Selecting condition.")
        # Input Condition
        try:
            # Click the Condition dropdown to display options
            condition_dropdown_xpath = "//label[@aria-label='Condition']"  # Adjust the XPath
            condition_dropdown = self.driver.find_element(By.XPATH, condition_dropdown_xpath)

            # Scroll the element into view
            self.driver.execute_script("arguments[0].scrollIntoView(true);", condition_dropdown)
            time.sleep(2)  # Wait for any overlay to disappear

            # Try clicking the dropdown
            try:
                condition_dropdown.click()
            except ElementClickInterceptedException:
                # Use JavaScript click as a fallback
                self.driver.execute_script("arguments[0].click();", condition_dropdown)

            time.sleep(5)  # Wait for options to be displayed

            # Find and click the desired option within specific div elements
            condition_choice_xpath = f"//div[@aria-selected='false']//span[contains(text(), '{condition}')]"
            condition_choice = self.driver.find_element(By.XPATH, condition_choice_xpath)
            
            self.driver.execute_script("arguments[0].scrollIntoView(true);", condition_choice)
            # Try clicking the dropdown
            try:
                condition_choice.click()
            except ElementClickInterceptedException:
                # Use JavaScript click as a fallback
                self.driver.execute_script("arguments[0].click();", condition_choice)

            print(f"Condition choice '{condition}' input successful.")
            logging.info(f"Condition choice '{condition}' input successful.")
            self.update_label(f"Condition choice input successful.")

            time.sleep(5)  # Wait for the selection to be processed
        except NoSuchWindowException as e:
            raise  # Reraise the exception to signal the failure
        except Exception as e:
            print(f"Error while selecting condition: {e}")
            logging.error(f"Error while selecting condition: {e}")
            self.update_label(f"Error while selecting condition.")
            traceback.print_exc()
            raise  # Reraise the exception to signal the failure

        time.sleep(random.uniform(1.0, 2.0))

    def input_description(self, description):
        if not description:
            raise Exception("No description provided.")
        
        message = """

        Notas sobre este producto:
        Este artículo es posible haya sido abierto para tomarle fotos para la publicación. En equipos como lámparas de techo, de pared, etc en ocasiones se evalúa que prendan y funcionen correctamente antes de proceder a venderlos.

        Ofertas Especiales:
        
        En el momento de la compra, pídeme un descuento para brindarte uno de los siguientes descuentos:
        - $5 de descuento en órdenes sobre $150 
        - $15 de descuento en órdenes sobre $200 
        - $30 de descuento en órdenes sobre $300

        Horarios y Ubicación:
        Estoy ubicado en Río Piedras, por la avenida Universidad. Generalmente, estoy disponible de 10 AM a 8 PM. Producto para recoger solamente por el momento.

        Más Productos:
        Para ver más productos que tengo a la venta, visita mi perfil de Facebook. En el primer post encontrarás el enlace a toda mi tienda con todos los productos disponibles.
        """

        description = "Detalles del producto:\n\n" + description + message

        print(f"Inputting description: {description[:600] + '...'}")
        logging.error(f"Inputting description.")
        self.update_label(f"Inputting description.")

        # Input Description 
        try:
            # Wait for the input field to be available
            wait = WebDriverWait (self.driver, 5)  # Adjust the timeout as needed
            # Adjusted XPath to target the specific title input field
            product_description_xpath = "//label[@aria-label='Description']//textarea"
            product_description = wait.until(EC.presence_of_element_located((By.XPATH, product_description_xpath)))


            # Insert text into the input field
            product_description.send_keys(description)
            print(f"Description input successful.")
            logging.info(f"Description input successful.")
            self.update_label(f"Description input successful.")

        except NoSuchWindowException as e:
            raise  # Reraise the exception to signal the failure
        except Exception as e:
            print(f"Error while inputting Description data: {e}")
            logging.error(f"Error while inputting Description data: {e}")
            self.update_label(f"Error while inputting Description data.")
            traceback.print_exc()  # This will print the full traceback
            raise  # Reraise the exception to signal the failure

    def adjust_description(self, description):
        if description.startswith("Título:"):
            first_newline_index = description.find('\n')
            if first_newline_index != -1:
                return description[first_newline_index + 1:].strip()
        return description

    def input_and_submit_tag(self, tag):        
        if not tag:
            raise Exception("No tag provided.")

        print(f"Inputting tag: {tag}")
        logging.info(f"Inputting tag.")
        self.update_label(f"Inputting tag.")

        try:

            # Focus the input field by clicking on it
            product_tags_input = self.driver.find_element(By.XPATH, "//label[@aria-label='Product tags']//textarea")
            product_tags_input.click()

            # Wait for 3 seconds to ensure the field becomes responsive
            time.sleep(1.5)  

            # Enter the tag
            product_tags_input.send_keys(tag)
            
            time.sleep(1.5)  # Adjust based on the responsiveness of application
            
            # Wait for the submit button to become clickable
            submit_button_xpath = "//div[@aria-label='Click to submit current value']"
            submit_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, submit_button_xpath))
            )
            
            # Click the submit button
            submit_button.click()

            print(f"Tag {tag} input successful.")
            logging.info(f"Tag input successful.")
            self.update_label(f"Tag input successful.")
        except NoSuchWindowException as e:
            raise  # Reraise the exception to signal the failure
        except Exception as e:
            print(f"Error inputting tag '{tag}': {e}")
            logging.error(f"Error inputting tag '{tag}': {e}")
            self.update_label(f"Error inputting tag.")
            raise  # Reraise the exception to signal the failure
            # Tags are optional inputs, no need to raise error unless the program is not running anymore.
            # TODO: Leave tag error raise temporarily while a bug is fixed with tag inputs.

    def input_product_tags(self, tags_string):
        tags = [tag.strip() for tag in tags_string.split(',') if tag.strip()]
        successful_tags = []
        skipped_tags = []

        for tag in tags:
            self.wait_for_resume()
            if len(successful_tags) >= 20:
                print("Reached the maximum tag limit. Skipping additional tags.")
                skipped_tags.extend(tags[len(successful_tags):])  # Add remaining tags to skipped list
                break  # Stop attempting to input more tags

            try:
                self.input_and_submit_tag(tag)
                successful_tags.append(tag)  # Add tag to successful list after input
            except ElementNotInteractableException:
                print(f"Skipped tag '{tag}': Element not interactable.")
                skipped_tags.append(tag)  # Add tag to skipped list
            except Exception as e:
                print(f"Error inputting tag '{tag}': {e}")
                skipped_tags.append(tag)  # Add tag to skipped list

        # Reporting
        if skipped_tags:
            print(f"Not all tags were inputted. {len(skipped_tags)} tags were skipped: {', '.join(skipped_tags)}")
        else:
            print("All tags were successfully inputted.")

    def input_sku(self, product_id, rack_id):
        if not product_id or not rack_id:
            raise Exception("No sku provided.")
    
        sku = f"{product_id} / {rack_id}"

        print(f"Inputting sku: {sku}")
        logging.error(f"Inputting sku.")
        self.update_label(f"Inputting sku.")
        # Input SKU/Product ID
        try:
            # Wait for the title input field to be available
            wait = WebDriverWait (self.driver, 10)  
            # Adjusted XPath to target the specific title input field
            sku_xpath = "//label[@aria-label='SKU']//input[@type='text']"

            sku_input = wait.until(EC.presence_of_element_located((By.XPATH, sku_xpath)))
            sku_input.send_keys(sku)  

            print(f"SKU {sku} input successful.")
            logging.info(f"SKU {sku} input successful.")
            self.update_label(f"SKU input successful.")
        except NoSuchWindowException as e:
            raise  # Reraise the exception to signal the failure
        except Exception as e:
            print(f"Error while inputting SKU data: {e}")
            logging.error(f"Error while inputting SKU data: {e}")
            self.update_label(f"Error while inputting SKU data.")
            traceback.print_exc()  # This will print the full traceback
            raise  # Reraise the exception to signal the failure

# Attempts to upload the product
    def click_next_button(self):
        print("Clicking the 'Next' button")
        logging.error(f"Clicking the 'Next' button")
        self.update_label(f"Clicking the 'Next' button")

        # Click the "Next" button
        try:
            # Use the aria-label to identify the "Next" button
            next_button_xpath = "//div[@aria-label='Next']"
            
            # Wait for the "Next" button to be clickable
            next_button = WebDriverWait (self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, next_button_xpath))
            )
            
            # Click the "Next" button
            next_button.click()
            print("Clicked on the 'Next' button.")
            logging.info(f"Clicked on the 'Next' button.")
            self.update_label(f"Clicked on the 'Next' button.")
        except NoSuchWindowException as e:
            raise  # Reraise the exception to signal the failure
        except Exception as e:
            print(f"Error clicking on 'Next' button: {e}")
            logging.error(f"Error clicking on 'Next' button: {e}")
            self.update_label(f"Error clicking on 'Next' button.")
            traceback.print_exc()
            raise  # Reraise the exception to signal the failure

        time.sleep(random.uniform(1.0, 2.0))


    def select_checkboxes(self):
        try:
            # Wait for checkboxes to be present
            checkboxes = WebDriverWait(self.driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//div[@data-visualcompletion='ignore-dynamic']"))
            )
            actions = ActionChains(self.driver)
            
            for checkbox in checkboxes:
                if checkbox.is_displayed() and checkbox.size['height'] > 0 and checkbox.size['width'] > 0:
                    # Scroll into view
                    self.driver.execute_script("arguments[0].scrollIntoView();", checkbox)
                    # Perform the click action
                    actions.move_to_element(checkbox).click().perform()
                    time.sleep(0.5)  # Optional: Adjust the sleep time based on your requirements
                else:
                    print(f"Checkbox not interactable: {checkbox}")

            print("All checkboxes with data-visualcompletion='ignore-dynamic' have been selected.")
            logging.info("All checkboxes with data-visualcompletion='ignore-dynamic' have been selected.")
        except Exception as e:
            print(f"Error selecting checkboxes: {e}")
            logging.error(f"Error selecting checkboxes: {e}")
            traceback.print_exc()


    def click_publish_button(self):
        print("Clicking the 'Publish' button")
        logging.error(f"Clicking the 'Publish' button")
        self.update_label(f"Clicking the 'Publish' button")

        # Code to click the "Publish" button
        try:
            # Use the aria-label to identify the "Publish" button
            publish_button_xpath = "//div[@aria-label='Publish']"
            
            # Wait for the "Publish" button to be clickable
            publish_button = WebDriverWait (self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, publish_button_xpath))
            )
            
            # Click the "Publish" button
            publish_button.click()
            print("Clicked on the 'Publish' button.")
            logging.info(f"Clicked on the 'Publish' button.")
            self.update_label(f"Clicked on the 'Publish' button.")
        except NoSuchWindowException as e:
            raise  # Reraise the exception to signal the failure
        except Exception as e:
            print(f"Error clicking on 'Publish' button: {e}")
            logging.error(f"Error clicking on 'Publish' button: {e}")
            self.update_label(f"Error clicking on 'Publish' button.")
            traceback.print_exc()
            raise  # Reraise the exception to signal the failure

        time.sleep(random.uniform(1.0, 2.0))

# Register product as uploaded in Excel
    def update_uploaded_status_in_excel(self, product_id, status='YES'):
        wb = load_workbook(self.excel_file_path)
        ws = wb[self.excel_sheet_name]

        # Find the column letters for "Product ID" and "Uploaded to Site"
        product_id_col = None
        uploaded_to_site_col = None

        for col in ws.iter_cols(1, ws.max_column, 1, 1):  # Iterate through the first row
            for cell in col:
                if cell.value == "Product ID":
                    product_id_col = cell.column_letter
                elif cell.value == "Uploaded to Site":
                    uploaded_to_site_col = cell.column_letter

        if not product_id_col or not uploaded_to_site_col:
            print("Required columns not found.")
            return

        # Update the "Uploaded to Site" status for the row with the matching "Product ID"
        for row in range(2, ws.max_row + 1):  # Skip the header row
            if ws[f'{product_id_col}{row}'].value == product_id:
                ws[f'{uploaded_to_site_col}{row}'] = status
                break  # Exit loop after updating

        wb.save(self.excel_file_path)
        print(f"Updated 'Uploaded to Site' status for Product ID: {product_id}")

# Creates a txt file in product folder to notify it has been uploaded 
    def create_uploaded_successful_txt(self, product_id):
        product_folder_path = None
        # Convert product_id to uppercase for case-insensitive comparison
        product_id_upper = product_id.upper()
        # Iterate through each folder in the product folder path
        for folder_name in os.listdir(self.product_folder_path):
            # Extract the product ID part from the folder name and convert to uppercase
            folder_product_id_upper = folder_name.split(' ', 1)[0].upper()
            if folder_product_id_upper == product_id_upper and os.path.isdir(os.path.join(self.product_folder_path, folder_name)):
                product_folder_path = os.path.join(self.product_folder_path, folder_name)
                break  # Found the matching folder, exit the loop

        if product_folder_path:
            today = datetime.today().strftime('%Y-%m-%d')
            filename = f"Uploaded to Marketplace on {today}.txt"
            filepath = os.path.join(product_folder_path, filename)
            with open(filepath, 'w') as f:
                f.write(f"Product ID {product_id} was successfully uploaded to Marketplace on {today}.")
            print(f"Created file {filename} in folder {product_folder_path}.")
        else:
            print(f"No folder for Product ID {product_id} was found.")


# After all products have been uploaded, enable inputs
    def on_process_complete(self):
        # Call enable_inputs to re-enable all inputs and the "Begin" button
        self.enable_inputs()
        self.begin_button.config(state=tk.NORMAL)  # Enable the "Begin" button after process completes
        self.pause_button.config(state=tk.DISABLED)  # Disable the "Pause" button after process completes
        self.resume_button.config(state=tk.DISABLED)  # Disable the "Resume" button after process completes


# When the program is closing, close browser and thread.
    def cleanup_resources(self):
        """Performs cleanup actions, such as quitting the WebDriver."""
        if self.driver:
            self.driver.quit()
            print("Browser closed.")
            logging.info("Browser closed.")
            self.driver = None

        # Ensure buttons are reset in case of abrupt stop
        self.on_process_complete()


    def on_close(self):
        """Handles the window close event."""
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.stop_event.set()  # Signal threads to stop

            print("Closing program...")
            logging.info("Closing program...")
            # Check if login_thread exists and is alive, then join it
            if hasattr(self, 'login_thread') and self.login_thread.is_alive():
                self.login_thread.join(timeout=5)  # Adjust the timeout as needed

            self.cleanup_resources()
            self.root.destroy()

            # Log the end of the program
            logging.info("Program ended\n")

# Start
if __name__ == "__main__":
    # Log the start of the program
    logging.info("Program started")

    uploader = MarketplaceUploader()
    uploader.run()

